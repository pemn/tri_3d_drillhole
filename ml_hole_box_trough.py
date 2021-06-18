#!python
# create a single linear hole downhole image from a sample box
# v1.0 06/2021 paulo.ernesto
'''
usage: $0 input_img#img*jpg,png trough_count=9 direction%horizontal,vertical output*xlsx display@
'''

import sys, os.path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import patches
from skimage import data, exposure, img_as_float
from skimage.io import imread
import skimage.morphology
import skimage.segmentation
import skimage.color
import skimage.util
import re
import openpyxl
import time
from PIL import Image as PILImage
from io import BytesIO

# import modules from a pyz (zip) file with same name as scripts
sys.path.insert(0, os.path.splitext(sys.argv[0])[0] + '.pyz')

from _gui import usage_gui, commalist

class ExcelImage(openpyxl.drawing.image.Image):
  def __init__(self, img_array):
    img = PILImage.fromarray(img_array)
    super().__init__(img)

  def _data(self):
    fp = BytesIO()
    self.ref.save(fp, format="png")
    fp.seek(0)
    return fp.read()

def filename_metadata(img_path):
  r = ['', 0, 0]
  m = re.search(r'(\w+)_(\d+\.\d+)_(\d+\.\d+)\.\w+$', img_path)
  if m:
    r = m.groups()
  return r
  
def crop_trough_data(img, t_list):
  r = []
  for t in t_list:
    c = [(t[0][0], img.shape[0] - t[1][0]), (t[0][1], img.shape[1] - t[1][1]), (0,0)]
    r.append(skimage.util.crop(np.copy(img), c))

  return r

def point_distance(p1, p2):
  dist = 0.0
  for i in range(len(p1)):
    dist += (p1[i] - p2[i]) ** 2.0
  dist = np.sqrt(dist)
  return(dist)

def split_core_trough(t_core, trough_count):
  ' split the mask containg the troughs into individual troughs '
  t_list = []
  bb = None
  for y in range(t_core.shape[0]):
    for x in range(t_core.shape[1]):
      if t_core[y, x]:
        if bb is None:
          bb = [[y, x], [y, x]]
        if y < bb[0][0]:
          bb[0][0] = y
        if x < bb[0][1]:
          bb[0][1] = x
        if y > bb[1][0]:
          bb[1][0] = y
        if x > bb[1][1]:
          bb[1][1] = x
  # print(bb)
  bb_w = np.subtract(bb[1], bb[0])
  # print("bb_w")
  # print(bb_w)
  bb0 = np.array(bb[0])
  bb_d = np.multiply(np.divide(1, trough_count), bb_w).astype(bb0.dtype)
  # print("bb_d")
  # print(bb_d)
  bb1 = np.add(bb0, bb_d)
  bb_s = np.multiply(bb_d, np.greater(trough_count, 1))
  # print("bb_s")
  # print(bb_s)
  for t in range(np.prod(trough_count)):
    # print("t", t)
    # print("bb")
    # print(bb0)
    # print(bb1)
    t_list.append([bb0.copy(), bb1.copy()])
    bb0 += bb_s
    bb1 += bb_s
  #print(t_list)
  #t_list = [bb]
  return t_list

def find_core_labels(labels, core = 0.35):
  ' create a mask on the sample box region containing the troughs '
  center = np.multiply(labels.shape, 0.5)
  nn_i = dict()
  nn_d = dict()
  for i in range(labels.shape[0]):
    for j in range(labels.shape[1]):
      l = labels[i,j]
      d = point_distance(center, (i,j))
      nn = nn_d.get(l)
      if nn is None or nn > d:
        nn_i[l] = (i,j)
        nn_d[l] = d
  core_dist = np.multiply(point_distance((0,0), center), core)
  m = np.zeros_like(labels)
  for k,v in nn_d.items():
    if v < core_dist:
      m |= labels == k

  return skimage.morphology.convex_hull_image(m)

def plot_ml_result(img, labels, t_list):
  fig, ax = plt.subplots(1, 2, sharex=True, sharey=True, figsize=(16,10))
  for i in range(len(ax)):  
    ax.flat[i].axis('off')

  ax.flat[0].imshow(labels / labels.max())
  ax.flat[1].imshow(img)
  for bb0, bb1 in t_list:
    rect = patches.Rectangle(np.flip(bb0), bb1[1] - bb0[1], bb1[0] - bb0[0], fill=False, edgecolor='green', linewidth=1)
    ax.flat[1].add_patch(rect)

def img_hole_box_trough(img_path, trough_setup, display):
  # Read the image
  img = imread(img_path)
  print(img.shape)

  labels = skimage.segmentation.slic(img, 16, 20.0, sigma=10, start_label=1)
  t_core = find_core_labels(labels)
  t_list = split_core_trough(t_core, trough_setup)

  o_list = crop_trough_data(img, t_list)
  if display:
    plot_ml_result(img, labels, t_list)
  return o_list

def ml_hole_box_trough(input_img, trough_count, direction, output, display):
  trough_count = int(trough_count)
  display = int(display)
  print("ml_hole_box_trough")

  trough_setup = [1, trough_count]
  if direction == 'horizontal':
    trough_setup = [trough_count, 1]

  input_img = commalist().parse(input_img).split()
  wb = openpyxl.Workbook()
  ws = wb.active
  ws['A1'] = 'img'
  ws['B1'] = 'hid'
  ws['C1'] = 'from'
  ws['D1'] = 'to'
  row = 2
  for img_path in input_img:
    metadata = filename_metadata(img_path)
    o_list = img_hole_box_trough(img_path, trough_setup, display)

    #o_list = [np.arange(256).reshape((16,16)), np.arange(256).reshape((16,16))]
    intervals = np.linspace(float(metadata[1]), float(metadata[2]), len(o_list) + 1, True)
    for n in range(len(o_list)):
      e_img = ExcelImage(o_list[n])
      ws.add_image(e_img, 'A%d' % (row))
      ws['B%d' % (row)] = metadata[0]
      ws['C%d' % (row)] = intervals[n]
      ws['D%d' % (row)] = intervals[n + 1]
      row += 1

  if output:
    wb.save(output)

  if display:
    plt.show()

main = ml_hole_box_trough

if __name__=="__main__":
  #main(*sys.argv[1:])
  usage_gui(__doc__)

# python ml_hole_box_trough.py SSC-FD00120/Arq_fotografico/SSC-FD00120_001_000.00_010.63.JPG
# [[152, 222], [2692,  459]]
